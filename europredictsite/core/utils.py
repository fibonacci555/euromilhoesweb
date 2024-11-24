import pandas as pd
import numpy as np
from openpyxl import load_workbook
from keras.models import Sequential
from keras.layers import LSTM, Dense, Bidirectional, Dropout
from keras.callbacks import EarlyStopping, CSVLogger
from keras.optimizers import RMSprop
from sklearn.model_selection import train_test_split
from keras.metrics import Precision, Recall

def predict():
    # Carregar o arquivo Excel gerado a partir do script de scraping
    wb = load_workbook("results.xlsx")
    ws = wb.active

    numeros = []

    # Validar e extrair os números do Excel
    for i in range(2, ws.max_row + 1):  # Começar da linha 2 para pular os cabeçalhos
        try:
            edicao = [
                int(ws[f'C{i}'].value),
                int(ws[f'D{i}'].value),
                int(ws[f'E{i}'].value),
                int(ws[f'F{i}'].value),
                int(ws[f'G{i}'].value),
                int(ws[f'H{i}'].value),
                int(ws[f'I{i}'].value)
            ]
            numeros.append(edicao)
        except (TypeError, ValueError):
            print(f"Dados inválidos na linha {i}, pulando esta linha.")
            continue

    # Converter a lista para um DataFrame
    df = pd.DataFrame(numeros, columns=['Num1', 'Num2', 'Num3', 'Num4', 'Num5', 'Star1', 'Star2'])
    print("Cabeçalho do DataFrame:")
    print(df.head())

    number_of_rows = df.shape[0]
    window_length = 10  # Número de sorteios anteriores a considerar

    # Função para criar vetor multi-hot
    def create_multi_hot(numbers, num_classes):
        multi_hot = np.zeros(num_classes)
        multi_hot[numbers - 1] = 1  # Subtrair 1 porque os índices começam em 0
        return multi_hot

    # Preparar os dados de treinamento
    X = []
    y_main = []
    y_stars = []

    for i in range(window_length, number_of_rows):
        # Usar os últimos 'window_length' sorteios como features
        past_draws = df.iloc[i - window_length:i]
        past_draws_flat = past_draws.values.flatten()
        X.append(past_draws_flat)

        # Criar labels multi-hot para o sorteio atual
        current_main_numbers = df.iloc[i][['Num1', 'Num2', 'Num3', 'Num4', 'Num5']].values
        current_star_numbers = df.iloc[i][['Star1', 'Star2']].values

        y_main.append(create_multi_hot(current_main_numbers, 50))
        y_stars.append(create_multi_hot(current_star_numbers, 12))

    X = np.array(X)
    y_main = np.array(y_main)
    y_stars = np.array(y_stars)

    # Dividir os dados em conjuntos de treinamento e validação
    X_train, X_val, y_main_train, y_main_val, y_stars_train, y_stars_val = train_test_split(
        X, y_main, y_stars, test_size=0.2, random_state=42
    )

    # Definir o modelo para os números principais
    def build_model(output_dim):
        model = Sequential()
        model.add(Bidirectional(LSTM(128, return_sequences=True), input_shape=(X_train.shape[1], 1)))
        model.add(Dropout(0.2))
        model.add(Bidirectional(LSTM(128)))
        model.add(Dropout(0.2))
        model.add(Dense(256, activation='relu'))
        model.add(Dense(output_dim, activation='sigmoid'))
        return model

    # Redimensionar X para ter a forma adequada para LSTM
    X_train_reshaped = X_train.reshape(X_train.shape[0], X_train.shape[1], 1)
    X_val_reshaped = X_val.reshape(X_val.shape[0], X_val.shape[1], 1)

    # Compilar o modelo para números principais
    model_main = build_model(50)
    model_main.compile(
        loss='binary_crossentropy',
        optimizer=RMSprop(learning_rate=0.001),
        metrics=['accuracy', Precision(), Recall()]
    )

    # Compilar o modelo para estrelas da sorte
    model_stars = build_model(12)
    model_stars.compile(
        loss='binary_crossentropy',
        optimizer=RMSprop(learning_rate=0.001),
        metrics=['accuracy', Precision(), Recall()]
    )

    # Callbacks para EarlyStopping e registro de logs
    early_stopping = EarlyStopping(patience=10, restore_best_weights=True)
    csv_logger_main = CSVLogger('training_log_main.csv', append=False)
    csv_logger_stars = CSVLogger('training_log_stars.csv', append=False)

    # Treinar o modelo para números principais
    model_main.fit(
        X_train_reshaped, y_main_train,
        epochs=100,
        batch_size=32,
        validation_data=(X_val_reshaped, y_main_val),
        callbacks=[early_stopping, csv_logger_main],
        verbose=1
    )

    # Treinar o modelo para estrelas da sorte
    model_stars.fit(
        X_train_reshaped, y_stars_train,
        epochs=100,
        batch_size=32,
        validation_data=(X_val_reshaped, y_stars_val),
        callbacks=[early_stopping, csv_logger_stars],
        verbose=1
    )

    # Preparar os dados para previsão (últimos 'window_length' sorteios)
    recent_draws = df.iloc[-window_length:]
    recent_draws_flat = recent_draws.values.flatten()
    to_predict = recent_draws_flat.reshape(1, -1, 1)

    # Fazer a previsão
    main_numbers_pred = model_main.predict(to_predict)
    star_numbers_pred = model_stars.predict(to_predict)

    # Converter previsões para números reais
    def get_numbers(predictions, num_numbers, num_classes):
        # Obter os índices dos maiores valores preditos
        indices = predictions[0].argsort()[-num_numbers:][::-1]
        # Adicionar 1 porque os números começam em 1
        numbers = indices + 1
        return np.sort(numbers)

    predicted_main_numbers = get_numbers(main_numbers_pred, 5, 50)
    predicted_star_numbers = get_numbers(star_numbers_pred, 2, 12)

    # Preparar a previsão final
    final_prediction = {
        'Números Principais': predicted_main_numbers.tolist(),
        'Estrelas da Sorte': predicted_star_numbers.tolist()
    }

    print("Números Previstos:")
    print(final_prediction)
    return final_prediction

if __name__ == "__main__":
    predict()
